import os, secrets, uuid, time, base64, html, io, json, sqlite3, threading, datetime
from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify, Response, send_file, make_response
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from geopy.distance import geodesic
from database import *
from dotenv import load_dotenv
import backup

load_dotenv()

app = Flask(__name__)
app.secret_key = os.getenv('SECRET_KEY', secrets.token_hex(32))
app.permanent_session_lifetime = datetime.timedelta(hours=8)
limiter = Limiter(app=app, key_func=get_remote_address, default_limits=["30/minute"])

@app.context_processor
def inject_globals():
    locked = False
    if session.get('role') == 'admin':
        locked = get_setting('system_locked', '0') == '1'
    return {'system_locked_global': locked}

@app.errorhandler(404)
def not_found(e):
    return render_template('errors/404.html'), 404

@app.errorhandler(403)
def forbidden(e):
    return render_template('errors/403.html'), 403

@app.errorhandler(500)
def server_error(e):
    return render_template('errors/500.html'), 500

FACILITY_LOCATION = (
    float(os.getenv('FACILITY_LAT', '25.8633556')),
    float(os.getenv('FACILITY_LNG', '43.4831471'))
)
ALLOWED_RADIUS = float(os.getenv('ALLOWED_RADIUS', '50'))

def login_required(f):
    from functools import wraps
    @wraps(f)
    def wrapper(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return wrapper

def admin_required(f):
    from functools import wraps
    @wraps(f)
    def wrapper(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        if session.get('role') != 'admin':
            flash('غير مصرح بالدخول', 'danger')
            return redirect(url_for('employee_dashboard'))
        return f(*args, **kwargs)
    return wrapper

def csrf_token():
    if 'csrf_token' not in session:
        session['csrf_token'] = secrets.token_hex(16)
    return session['csrf_token']

app.jinja_env.globals.update(csrf_token=csrf_token, get_user_by_id=get_user_by_id)

@app.context_processor
def inject_globals():
    return {
        'now': ksa().strftime('%H:%M'),
        'facility_lat': FACILITY_LOCATION[0],
        'facility_lng': FACILITY_LOCATION[1],
        'allowed_radius': ALLOWED_RADIUS,
    }

def build_report(date_str):
    amap     = get_report(date_str)
    soldiers = get_soldiers()
    present = absent = on_leave = no_out = 0
    rows = []
    sids = [s[0] for s in soldiers]
    leaves_map = bulk_get_leaves(sids)
    for sid, nm, _, _, rank, phone in soldiers:
        rank_s = rank or ''
        nm_full = f"{rank_s} / {nm}" if rank_s else nm
        lv = leaves_map.get(sid)
        note_raw = amap[sid]['note'] if sid in amap and amap[sid]['note'] else ''
        if lv:
            rows.append({
                'name': nm_full, 'status': 'إجازة', 'note': '',
                'leave_label': lv['label'], 'leave_end': lv['end']
            })
            on_leave += 1
        elif sid in amap:
            ci = amap[sid]['check_in'] or '---'
            co = amap[sid]['check_out'] or '---'
            loc_in = amap[sid].get('loc','')
            loc_out = amap[sid].get('loc_out','')
            if amap[sid]['check_in'] and not amap[sid]['check_out']:
                rows.append({
                    'name': nm_full, 'status': 'بدون خروج', 'note': note_raw,
                    'check_in': ci, 'check_out': '---',
                    'loc_in': loc_in, 'loc_out': ''
                })
                no_out += 1
            else:
                rows.append({
                    'name': nm_full, 'status': 'مكتمل', 'note': note_raw,
                    'check_in': ci, 'check_out': co,
                    'loc_in': loc_in, 'loc_out': loc_out
                })
            present += 1
        else:
            rows.append({
                'name': nm_full, 'status': 'غائب', 'note': '',
                'check_in': '---', 'check_out': '---'
            })
            absent += 1
    return {'present': present, 'absent': absent, 'on_leave': on_leave, 'no_out': no_out, 'rows': rows}

# ── Auth ──

@app.route('/login', methods=['GET', 'POST'])
@limiter.limit("10/minute")
def login():
    if request.method == 'POST':
        username = request.form.get('username', '').strip().lower()
        password = request.form.get('password', '')
        user = get_user(username)
        if not user or not check_password(user['password'], password):
            flash('بيانات الدخول غير صحيحة', 'danger')
            return render_template('login.html')
        if user['is_blocked']:
            flash('هذا الحساب موقوف. تواصل مع القائد.', 'danger')
            return render_template('login.html')
        device_token = request.cookies.get('device_token', '')
        if not device_token:
            device_token = secrets.token_hex(32)
        if user['device_uid']:
            stored_tokens = [d.strip() for d in user['device_uid'].split(',') if d.strip()]
            if device_token not in stored_tokens:
                if len(stored_tokens) >= 2:
                    session.clear()
                    flash('تم تجاوز الحد المسموح به من الأجهزة (جهازان كحد أقصى). للدعم تواصل مع الإدارة.', 'danger')
                    return render_template('login.html')
                stored_tokens.append(device_token)
                db_run("UPDATE users SET device_uid=? WHERE id=?", (','.join(stored_tokens), user['id']))
                admins = get_admin_ids()
                for (aid,) in admins:
                    push_notif(aid, f"⚠️ تنبيه: جهاز جديد لحساب {user['full_name']}")
        else:
            db_run("UPDATE users SET device_uid=? WHERE id=?", (device_token, user['id']))
        if not user['password'].startswith('$2'):
            db_run("UPDATE users SET password=? WHERE id=?", (hash_password(password), user['id']))
        session.clear()
        session['user_id'] = user['id']
        session['username'] = user['username']
        session['full_name'] = user['full_name']
        session['role'] = user['role']
        session['user_avatar'] = user.get('avatar', '')
        session.permanent = True
        redirect_url = url_for('admin_dashboard' if user['role'] == 'admin' else 'employee_dashboard')
        resp = make_response(render_template('login_loading.html', redirect_url=redirect_url))
        resp.set_cookie('device_token', device_token, max_age=365*24*3600, httponly=True, samesite='Lax')
        return resp
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

@app.route('/')
def index():
    if 'user_id' in session:
        if session.get('role') == 'admin':
            return redirect(url_for('admin_dashboard'))
        return redirect(url_for('employee_dashboard'))
    return redirect(url_for('login'))

@app.route('/profile', methods=['GET', 'POST'])
@login_required
def profile():
    uid = session['user_id']
    user = get_user_by_id(uid)
    if request.method == 'POST':
        if request.form.get('_csrf', '') != session.get('csrf_token', ''):
            flash('خطأ في التحقق', 'danger'); return redirect(url_for('profile'))
        if 'avatar' in request.files:
            f = request.files['avatar']
            if f and f.filename:
                import base64
                raw = f.read()
                ext = f.filename.rsplit('.',1)[-1].lower() if '.' in f.filename else 'png'
                if ext not in ('png','jpg','jpeg','gif','webp'): ext = 'png'
                b64 = base64.b64encode(raw).decode()
                avatar = f"data:image/{ext};base64,{b64}"
                db_run("UPDATE users SET avatar=? WHERE id=?", (avatar, uid))
                session['user_avatar'] = avatar
                flash('تم تغيير الصورة', 'success')
        return redirect(url_for('profile'))
    return render_template('profile.html', user=user)

# ── Admin ──
@app.route('/admin/system-lock', methods=['POST'])
@admin_required
def admin_system_lock():
    if request.form.get('_csrf', '') != session.get('csrf_token', ''):
        flash('خطأ في التحقق', 'danger'); return redirect(url_for('admin_dashboard'))
    cur = int(get_setting('system_locked', '0'))
    new = '0' if cur else '1'
    set_setting('system_locked', new)
    log_action(session['user_id'], 'system_lock', None, f'قفل النظام: {"مقفل" if new=="1" else "مفتوح"}')
    flash('تم ' + ('قفل' if new=='1' else 'فتح') + ' النظام', 'success')
    _auto_save(); return redirect(url_for('admin_dashboard'))

@app.route('/admin')
@admin_required
def admin_dashboard():
    dt = ksa().strftime('%Y-%m-%d')
    rpt = build_report(dt)
    soldiers = get_soldiers()
    pending = len(get_pending_requests())
    unread = len(get_unread(session['user_id']))
    system_locked = get_setting('system_locked', '0') == '1'
    return render_template('admin/dashboard.html', rpt=rpt, total=len(soldiers),
        pending_req=pending, unread=unread, system_locked=system_locked)

@app.route('/admin/users')
@admin_required
def admin_users():
    soldiers = get_soldiers()
    device_map = {}
    for sid, _, _, _, _, _ in soldiers:
        u = get_user_by_id(sid)
        if u and u.get('device_uid'):
            device_map[sid] = True
    return render_template('admin/users.html', soldiers=soldiers, user_device_map=device_map)

@app.route('/admin/users/add', methods=['GET','POST'])
@admin_required
def admin_add_user():
    if request.method == 'POST':
        if request.form.get('_csrf', '') != session.get('csrf_token', ''):
            flash('خطأ في التحقق', 'danger'); return render_template('admin/user_form.html', user=None)
        name = request.form.get('full_name', '').strip()
        uname = request.form.get('username', '').strip().lower()
        password = request.form.get('password', '').strip()
        rank = request.form.get('rank_title', '').strip()
        phone = request.form.get('phone_number', '').strip()
        if not name or not uname or not password:
            flash('جميع الحقول المطلوبة يجب أن تُملأ', 'danger')
            return render_template('admin/user_form.html', user=None)
        ok, msg = add_user(uname, password, name, 'employee', rank, phone)
        flash(msg, 'success' if ok else 'danger')
        if ok: _auto_save(); return redirect(url_for('admin_users'))
    return render_template('admin/user_form.html', user=None)

@app.route('/admin/users/<int:uid>/edit', methods=['GET','POST'])
@admin_required
def admin_edit_user(uid):
    user = get_user_by_id(uid)
    if not user:
        flash('المستخدم غير موجود', 'danger')
        return redirect(url_for('admin_users'))
    if request.method == 'POST':
        if request.form.get('_csrf', '') != session.get('csrf_token', ''):
            flash('خطأ في التحقق', 'danger'); return redirect(url_for('admin_users'))
        fields = {'full_name':'full_name','username':'username','rank_title':'rank_title','phone_number':'phone_number'}
        for fld, col in fields.items():
            val = request.form.get(fld, '').strip()
            if val and val != user.get(col):
                if fld == 'username':
                    update_user(uid, 'username', val)
                else:
                    db_run(f"UPDATE users SET {col}=? WHERE id=?", (val, uid))
        pw = request.form.get('password', '').strip()
        if pw:
            update_user(uid, 'password', pw)
        _auto_save(); flash('تم التحديث بنجاح', 'success')
        return redirect(url_for('admin_users'))
    return render_template('admin/user_form.html', user=user)

@app.route('/admin/users/<int:uid>/delete', methods=['POST'])
@admin_required
def admin_delete_user(uid):
    if request.form.get('_csrf', '') != session.get('csrf_token', ''):
        flash('خطأ في التحقق', 'danger'); return redirect(url_for('admin_users'))
    u = get_user_by_id(uid)
    delete_user(uid)
    log_action(session['user_id'], 'soft_delete_user', uid, f'حذف المستخدم: {u["full_name"] if u else uid}')
    _auto_save(); flash('تم الحذف', 'success')
    return redirect(url_for('admin_users'))

@app.route('/admin/users/<int:uid>/restore', methods=['POST'])
@admin_required
def admin_restore_user(uid):
    if request.form.get('_csrf', '') != session.get('csrf_token', ''):
        flash('خطأ في التحقق', 'danger'); return redirect(url_for('admin_users'))
    restore_user(uid)
    log_action(session['user_id'], 'restore_user', uid, f'استعادة المستخدم: {uid}')
    flash('تمت الاستعادة', 'success')
    return redirect(url_for('admin_users'))

@app.route('/admin/deleted-users')
@admin_required
def admin_deleted_users():
    soldiers = get_deleted_soldiers()
    return render_template('admin/deleted_users.html', soldiers=soldiers)

@app.route('/admin/users/<int:uid>/toggle-block', methods=['POST'])
@admin_required
def admin_toggle_block(uid):
    if request.form.get('_csrf', '') != session.get('csrf_token', ''):
        flash('خطأ في التحقق', 'danger'); return redirect(url_for('admin_users'))
    result = toggle_block(uid)
    if result is not None:
        u = get_user_by_id(uid)
        log_action(session['user_id'], 'toggle_block', uid, f'إيقاف المستخدم: {u["full_name"] if u else uid}')
        _auto_save(); flash('تم التفعيل' if result == 0 else 'تم الإيقاف', 'success')
    return redirect(url_for('admin_users'))

@app.route('/admin/self/reset-device', methods=['POST'])
@admin_required
def admin_reset_self_device():
    if request.form.get('_csrf', '') != session.get('csrf_token', ''):
        flash('خطأ في التحقق', 'danger')
        return redirect(request.referrer or url_for('admin_dashboard'))
    reset_device_uid(session['user_id'])
    flash('تم إعادة تعيين أجهزتك. سجل دخول من الجهاز الجديد.', 'success')
    return redirect(url_for('admin_dashboard'))

@app.route('/admin/users/<int:uid>/reset-device', methods=['POST'])
@admin_required
def admin_reset_device(uid):
    if request.form.get('_csrf', '') != session.get('csrf_token', ''):
        flash('خطأ في التحقق', 'danger'); return redirect(url_for('admin_users'))
    reset_device_uid(uid)
    user = get_user_by_id(uid)
    log_action(session['user_id'], 'reset_device', uid, f'إعادة تعيين جهاز: {user["full_name"] if user else uid}')
    flash(f'تم إعادة تعيين الجهاز لـ {user["full_name"]}', 'success')
    return redirect(url_for('admin_users'))

@app.route('/admin/users/search', methods=['POST'])
@admin_required
def admin_search_user():
    q = request.form.get('query', '').strip().lower()
    user = get_user(q)
    if user and user['role'] == 'employee':
        lv = get_leave(user['id'])
        hist = get_history_records(user['id'])
        loc_data = db_get("SELECT latitude,longitude,timestamp,action FROM attendance WHERE user_id=? ORDER BY id DESC LIMIT 1", (user['id'],))
        reqs = db_get("SELECT duration_label,request_date,status FROM leave_requests WHERE user_id=? ORDER BY id DESC", (user['id'],))
        return render_template('admin/user_search_result.html', u=user, lv=lv, hist=hist, loc=loc_data, reqs=reqs)
    flash('لم يتم العثور على العسكري', 'warning')
    return redirect(url_for('admin_users'))

@app.route('/admin/attendance/manual', methods=['GET','POST'])
@admin_required
def admin_manual_attendance():
    soldiers = get_soldiers()
    if request.method == 'POST':
        if request.form.get('_csrf', '') != session.get('csrf_token', ''):
            flash('خطأ في التحقق', 'danger'); return redirect(url_for('admin_manual_attendance'))
        uid = int(request.form.get('user_id', 0))
        action = request.form.get('action', '')
        note = request.form.get('note', '')
        if uid and action in ('check_in','check_out'):
            ok, err = record_attendance(uid, action, 0, 0, f"تسجيل يدوي من القائد - {note}" if note else "تسجيل يدوي من القائد")
            u = get_user_by_id(uid)
            act_ar = "دخول" if action=='check_in' else "خروج"
            if not ok:
                flash(f'لم يتم تسجيل {act_ar} لـ {u["full_name"]}: {err}', 'warning')
            else:
                push_notif(uid, f"تم تسجيل {act_ar} يدوياً من قِبل القائد.")
                _auto_save(); flash(f'تم تسجيل {act_ar} لـ {u["full_name"]}', 'success')
        return redirect(url_for('admin_manual_attendance'))
    return render_template('admin/manual_att.html', soldiers=soldiers)

@app.route('/admin/reports')
@admin_required
def admin_reports():
    dt = ksa().strftime('%Y-%m-%d')
    rpt = build_report(dt)
    dates = get_dates()
    return render_template('admin/reports.html', rpt=rpt, dt=dt, dates=dates, now=ksa().strftime('%H:%M'))

@app.route('/admin/reports/<date_str>')
@admin_required
def admin_report_date(date_str):
    rpt = build_report(date_str)
    return render_template('admin/reports.html', rpt=rpt, dt=date_str, dates=get_dates(), now=ksa().strftime('%H:%M'))

@app.route('/admin/reports/<date_str>/export')
@admin_required
def admin_export_excel(date_str):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    wb = Workbook()
    ws = wb.active
    ws.title = f"تقرير {date_str}"
    ws.sheet_view.rightToLeft = True
    headers = ['الاسم', 'الرتبة', 'دخول', 'خروج', 'الحالة', 'ملاحظة']
    hdr_font = Font(name='Tajawal', bold=True, size=12, color='FFFFFF')
    hdr_fill = PatternFill('solid', fgColor='2EA043')
    thin = Side(style='thin', color='30363D')
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)
    soldiers = get_soldiers()
    amap = get_report(date_str)
    sids = [s[0] for s in soldiers]
    leaves_map = bulk_get_leaves(sids)
    row = 2
    body_font = Font(name='Tajawal', size=11)
    for sid, nm, _, _, rank, phone in soldiers:
        nm_full = f"{rank} / {nm}" if rank else nm
        lv = leaves_map.get(sid)
        if lv:
            ci = '---'; co = '---'; note = f"إجازة: {lv['label']}"; status = 'إجازة'
        elif sid in amap:
            ci = amap[sid]['check_in'] or '---'
            co = amap[sid]['check_out'] or '---'
            note = amap[sid]['note'] or ''
            status = 'بدون خروج' if amap[sid]['check_in'] and not amap[sid]['check_out'] else 'مكتمل'
        else:
            ci = '---'; co = '---'; note = ''; status = 'غائب'
        vals = [nm_full, rank or '---', ci, co, status, note]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=c, value=v)
            cell.font = body_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)
        row += 1
    ws.column_dimensions['A'].width = 30
    for col in 'BCDEF': ws.column_dimensions[col].width = 16
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return Response(out.getvalue(), mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename=report_{date_str}.xlsx'})

@app.route('/admin/leaves')
@admin_required
def admin_leaves():
    active = get_active_leaves()
    requests_pending = get_pending_requests()
    soldiers = get_soldiers()
    return render_template('admin/leaves.html', active=active, requests=requests_pending, soldiers=soldiers)

@app.route('/admin/leaves/register', methods=['POST'])
@admin_required
def admin_register_leave():
    if request.form.get('_csrf', '') != session.get('csrf_token', ''):
        flash('خطأ في التحقق', 'danger'); return redirect(url_for('admin_leaves'))
    uid = int(request.form.get('user_id', 0))
    hours = int(request.form.get('hours', 0))
    if uid and hours > 0:
        std, etd = set_leave(uid, hours)
        u = get_user_by_id(uid)
        log_action(session['user_id'], 'register_leave', uid, f'رخصة {hours} ساعة')
        push_notif(uid, f"تم تسجيل رخصة {hours} ساعة. تنتهي: {etd.strftime('%Y-%m-%d %H:%M')}")
        _auto_save(); flash(f'رخصة {hours} ساعة لـ {u["full_name"]}', 'success')
    return redirect(url_for('admin_leaves'))

@app.route('/admin/leaves/<int:uid>/revoke', methods=['POST'])
@admin_required
def admin_revoke_leave(uid):
    if request.form.get('_csrf', '') != session.get('csrf_token', ''):
        flash('خطأ في التحقق', 'danger'); return redirect(url_for('admin_leaves'))
    revoke_leave(uid)
    u = get_user_by_id(uid)
    log_action(session['user_id'], 'revoke_leave', uid, f'سحب رخصة: {u["full_name"] if u else uid}')
    push_notif(uid, "تم سحب رخصتك وإعادتك للواجب الميداني.")
    _auto_save(); flash(f'تم سحب رخصة {u["full_name"]}', 'success')
    return redirect(url_for('admin_leaves'))

@app.route('/admin/leave-requests/<int:rid>/approve', methods=['POST'])
@admin_required
def admin_approve_request(rid):
    if request.form.get('_csrf', '') != session.get('csrf_token', ''):
        flash('خطأ في التحقق', 'danger'); return redirect(url_for('admin_leaves'))
    result = approve_leave_request(rid)
    if result:
        uid, hours = result
        u = get_user_by_id(uid)
        _, etd = set_leave(uid, hours)
        log_action(session['user_id'], 'approve_leave', uid, f'اعتماد رخصة {hours} ساعة')
        push_notif(uid, f"تم اعتماد رخصتك {hours} ساعة. تنتهي: {etd.strftime('%Y-%m-%d %H:%M')}")
        _auto_save(); flash('تم اعتماد الرخصة', 'success')
    return redirect(url_for('admin_leaves'))

@app.route('/admin/leave-requests/<int:rid>/reject', methods=['POST'])
@admin_required
def admin_reject_request(rid):
    if request.form.get('_csrf', '') != session.get('csrf_token', ''):
        flash('خطأ في التحقق', 'danger'); return redirect(url_for('admin_leaves'))
    uid = reject_leave_request(rid)
    if uid:
        log_action(session['user_id'], 'reject_leave', uid, 'رفض طلب رخصة')
        push_notif(uid, "تم رفض طلب الرخصة.")
    _auto_save(); flash('تم رفض الطلب', 'success')
    return redirect(url_for('admin_leaves'))

@app.route('/admin/shifts')
@admin_required
def admin_shifts():
    today = get_today_shifts()
    archive = get_shift_archive()
    soldiers = get_soldiers()
    return render_template('admin/shifts.html', today=today, archive=archive, soldiers=soldiers)

@app.route('/admin/shifts/set', methods=['POST'])
@admin_required
def admin_set_shift():
    if request.form.get('_csrf', '') != session.get('csrf_token', ''):
        flash('خطأ في التحقق', 'danger'); return redirect(url_for('admin_shifts'))
    duty = request.form.get('full_name', '').strip()
    if duty:
        set_shift(duty)
        _auto_save(); flash(f'تم تسجيل المستلم: {duty}', 'success')
    return redirect(url_for('admin_shifts'))

@app.route('/admin/shifts/<int:sid>/delete', methods=['POST'])
@admin_required
def admin_delete_shift(sid):
    if request.form.get('_csrf', '') != session.get('csrf_token', ''):
        flash('خطأ في التحقق', 'danger'); return redirect(url_for('admin_shifts'))
    delete_shift(sid)
    _auto_save(); flash('تم فك الاستلامة', 'success')
    return redirect(url_for('admin_shifts'))

@app.route('/admin/security')
@admin_required
def admin_security():
    alerts = get_security_alerts()
    hist = get_history_records()
    soldiers = get_soldiers()
    return render_template('admin/security.html', alerts=alerts, history=hist, soldiers=soldiers)

@app.route('/admin/security/add-history', methods=['POST'])
@admin_required
def admin_add_history():
    if request.form.get('_csrf', '') != session.get('csrf_token', ''):
        flash('خطأ في التحقق', 'danger'); return redirect(url_for('admin_security'))
    uid = int(request.form.get('user_id', 0))
    note = request.form.get('note', '').strip()
    if uid and note:
        add_history(uid, note)
        log_action(session['user_id'], 'add_history', uid, f'تسجيل سابقة: {note[:100]}')
        push_notif(uid, f"تم تسجيل ملاحظة على حسابك من قِبل القيادة:\n{note}")
        flash('تم تسجيل الملاحظة', 'success')
    return redirect(url_for('admin_security'))

@app.route('/admin/security/del-history/<int:hid>', methods=['POST'])
@admin_required
def admin_delete_history(hid):
    if request.form.get('_csrf', '') != session.get('csrf_token', ''):
        flash('خطأ في التحقق', 'danger'); return redirect(url_for('admin_security'))
    delete_history(hid)
    log_action(session['user_id'], 'delete_history', hid, 'حذف سابقة')
    flash('تم حذف السابقة', 'success')
    return redirect(url_for('admin_security'))

@app.route('/admin/security/reset-device/<int:uid>', methods=['POST'])
@admin_required
def admin_security_reset_device(uid):
    if request.form.get('_csrf', '') != session.get('csrf_token', ''):
        flash('خطأ في التحقق', 'danger'); return redirect(url_for('admin_security'))
    reset_device_uid(uid)
    u = get_user_by_id(uid)
    log_action(session['user_id'], 'security_reset_device', uid, f'فك ربط جهاز: {u["full_name"] if u else uid}')
    flash('تم فك ربط الجهاز', 'success')
    return redirect(url_for('admin_security'))

@app.route('/admin/security/dismiss-alert/<int:aid>', methods=['POST'])
@admin_required
def admin_dismiss_alert(aid):
    db_run("DELETE FROM security_alerts WHERE id=?", (aid,))
    log_action(session['user_id'], 'dismiss_alert', aid, 'حذف تنبيه أمني')
    flash('تم حذف التنبيه', 'success')
    return redirect(url_for('admin_security'))

@app.route('/admin/security/test-alert', methods=['POST'])
@admin_required
def admin_test_alert():
    if request.form.get('_csrf', '') != session.get('csrf_token', ''):
        flash('خطأ في التحقق', 'danger'); return redirect(url_for('admin_security'))
    add_security_alert(session['user_id'], 'اختبار', 'تم إنشاء هذا التنبيه يدوياً للتأكد من عمل النظام.')
    admins = get_admin_ids()
    for (aid,) in admins:
        push_notif(aid, '✅ اختبار: نظام التنبيهات الأمنية يعمل بشكل صحيح')
    flash('✅ تم إنشاء تنبيه اختباري', 'success')
    return redirect(url_for('admin_security'))

@app.route('/admin/circulars')
@admin_required
def admin_circulars():
    circular = get_circular()
    return render_template('admin/circulars.html', circular=circular)

@app.route('/admin/circulars/set', methods=['POST'])
@admin_required
def admin_set_circular():
    if request.form.get('_csrf', '') != session.get('csrf_token', ''):
        flash('خطأ في التحقق', 'danger'); return redirect(url_for('admin_circulars'))
    content = request.form.get('content', '').strip()
    if content:
        set_circular(content)
        flash('تم حفظ التعميم', 'success')
    return redirect(url_for('admin_circulars'))

@app.route('/admin/circulars/broadcast', methods=['POST'])
@admin_required
def admin_broadcast():
    if request.form.get('_csrf', '') != session.get('csrf_token', ''):
        flash('خطأ في التحقق', 'danger'); return redirect(url_for('admin_circulars'))
    content = request.form.get('content', '').strip()
    if content:
        set_circular(content)
        employees = db_get("SELECT id FROM users WHERE role='employee' AND is_blocked=0")
        for eid, in employees:
            push_notif(eid, f"📢 تعميم جديد: {content[:100]}{'...' if len(content)>100 else ''}")
        flash('تم بث التعميم لجميع الأفراد', 'success')
    return redirect(url_for('admin_circulars'))

@app.route('/admin/reports-panel')
@admin_required
def admin_reports_panel():
    reports = get_open_reports()
    report_data = []
    for rid, sid, txt, cat in reports:
        u = get_user_by_id(sid)
        report_data.append({'id': rid, 'sender': u['full_name'] if u else sid, 'text': txt, 'date': cat})
    return render_template('admin/reports_panel.html', reports=report_data)

@app.route('/admin/reports/<int:rid>/reply', methods=['POST'])
@admin_required
def admin_reply_report(rid):
    if request.form.get('_csrf', '') != session.get('csrf_token', ''):
        flash('خطأ في التحقق', 'danger'); return redirect(url_for('admin_reports_panel'))
    reply = request.form.get('reply', '').strip()
    if reply:
        uid = reply_report(rid, reply)
        if uid:
            push_notif(uid, f"رد القيادة على بلاغك:\n{reply}")
        flash('تم إرسال الرد', 'success')
    return redirect(url_for('admin_reports_panel'))

@app.route('/admin/notifications')
@admin_required
def admin_notifications():
    notifs = get_all_notifications(session['user_id'])
    unread_ids = [n[0] for n in notifs if not n[3]]
    if unread_ids:
        mark_read(session['user_id'], unread_ids)
    return render_template('admin/notifications.html', notifications=notifs)

@app.route('/admin/notifications/clear', methods=['POST'])
@admin_required
def admin_clear_read_notifications():
    if request.form.get('_csrf', '') != session.get('csrf_token', ''):
        flash('خطأ في التحقق', 'danger'); return redirect(url_for('admin_notifications'))
    delete_read_notifications(session['user_id'])
    flash('تم حذف الإشعارات المقروءة', 'success')
    return redirect(url_for('admin_notifications'))

@app.route('/admin/notifications/delete-all', methods=['POST'])
@admin_required
def admin_delete_all_notifications():
    if request.form.get('_csrf', '') != session.get('csrf_token', ''):
        flash('خطأ في التحقق', 'danger'); return redirect(url_for('admin_notifications'))
    delete_unread_notifications(session['user_id'])
    flash('تم حذف جميع الإشعارات', 'success')
    return redirect(url_for('admin_notifications'))

@app.route('/admin/send-message', methods=['GET','POST'])
@admin_required
def admin_send_message():
    soldiers = get_soldiers()
    if request.method == 'POST':
        uid = int(request.form.get('user_id', 0))
        msg = request.form.get('message', '').strip()
        if uid and msg:
            push_notif(uid, f"رسالة من القيادة:\n{msg}")
            u = get_user_by_id(uid)
            flash(f'تم إرسال الرسالة لـ {u["full_name"]}', 'success')
        return redirect(url_for('admin_send_message'))
    return render_template('admin/send_message.html', soldiers=soldiers)

# ── Employee ──
@app.route('/employee')
@login_required
def employee_dashboard():
    if session.get('role') != 'employee':
        return redirect(url_for('admin_dashboard'))
    user = get_user_by_id(session['user_id'])
    lv = get_leave(user['id'])
    circular = get_circular()
    unread = len(get_unread(user['id']))
    today_att = db_get("SELECT action,timestamp FROM attendance WHERE user_id=? AND DATE(timestamp)=? ORDER BY id", (user['id'], ksa().strftime('%Y-%m-%d')))
    return render_template('employee/dashboard.html', user=user, lv=lv, circular=circular, unread=unread, attendance=today_att)

@app.route('/employee/attendance')
@login_required
def employee_attendance():
    today = get_today_attendance(session['user_id'])
    return render_template('employee/attendance.html', has_check_in='check_in' in today, has_check_out='check_out' in today)

@app.route('/employee/check-in', methods=['POST'])
@login_required
def employee_check_in():
    return handle_attendance('check_in')

@app.route('/employee/check-out', methods=['POST'])
@login_required
def employee_check_out():
    return handle_attendance('check_out')

def handle_attendance(action):
    user_id = session['user_id']
    if get_setting('system_locked', '0') == '1':
        return jsonify({'ok': False, 'msg': 'النظام مقفل من قبل الإدارة'}), 403
    now = time.time()
    last = session.get('last_att', 0)
    if now - last < 3:
        return jsonify({'ok': False, 'msg': 'الرجاء الانتظار 3 ثوانٍ بين كل محاولة'}), 429
    data = request.get_json()
    if not data:
        return jsonify({'ok': False, 'msg': 'بيانات غير صالحة'}), 400
    lat = data.get('latitude')
    lng = data.get('longitude')
    if lat is None or lng is None:
        return jsonify({'ok': False, 'msg': 'الرجاء مشاركة الموقع'}), 400
    dist = geodesic(FACILITY_LOCATION, (lat, lng)).meters
    if dist > ALLOWED_RADIUS:
        return jsonify({'ok': False, 'msg': f'أنت خارج النطاق المسموح. مسافتك: {dist:.0f}م (الحد: {ALLOWED_RADIUS}م)'}), 400
    ok, err = record_attendance(user_id, action, lat, lng, note)
    if not ok:
        return jsonify({'ok': False, 'msg': err}), 409
    session['last_att'] = now
    if os.getenv('GH_TOKEN'): threading.Thread(target=backup.do_backup, daemon=True).start()
    verb = "تم تسجيل الحضور ✅" if action == 'check_in' else "تم تسجيل الانصراف 🔴"
    return jsonify({'ok': True, 'msg': f'{verb}\nالمسافة: {dist:.0f}م'})

@app.route('/employee/leave')
@login_required
def employee_leave():
    user = get_user_by_id(session['user_id'])
    lv = get_leave(user['id'])
    reqs = db_get("SELECT duration_label,request_date,status FROM leave_requests WHERE user_id=? ORDER BY id DESC LIMIT 5", (user['id'],))
    return render_template('employee/leave.html', lv=lv, requests=reqs)

@app.route('/employee/leave/request', methods=['POST'])
@login_required
def employee_request_leave():
    if request.form.get('_csrf', '') != session.get('csrf_token', ''):
        flash('خطأ في التحقق', 'danger'); return redirect(url_for('employee_leave'))
    hours = int(request.form.get('hours', 0))
    if hours in (24, 48, 72):
        request_leave(session['user_id'], hours)
        _auto_save(); flash('تم رفع طلب الرخصة. بانتظار موافقة القائد.', 'success')
    return redirect(url_for('employee_leave'))

@app.route('/employee/report', methods=['GET','POST'])
@login_required
def employee_report():
    if request.method == 'POST':
        text = request.form.get('report_text', '').strip()
        if text:
            send_report(session['user_id'], text)
            flash('تم رفع البلاغ للقيادة', 'success')
            return redirect(url_for('employee_dashboard'))
    return render_template('employee/report.html')

@app.route('/employee/message-leader', methods=['GET','POST'])
@login_required
def employee_message_leader():
    if request.method == 'POST':
        text = request.form.get('message', '').strip()
        if text:
            user = get_user_by_id(session['user_id'])
            admins = get_admin_ids()
            for (aid,) in admins:
                push_notif(aid, f"رسالة من {user['full_name']}:\n{text}")
            flash('تم إرسال رسالتك للقيادة', 'success')
            return redirect(url_for('employee_dashboard'))
    return render_template('employee/message_leader.html')

@app.route('/employee/notifications')
@login_required
def employee_notifications():
    notifs = get_all_notifications(session['user_id'])
    unread_ids = [n[0] for n in notifs if not n[3]]
    if unread_ids:
        mark_read(session['user_id'], unread_ids)
    return render_template('employee/notifications.html', notifications=notifs)

# ── Backup & Restore ──
def _backup_rows(table):
    try:
        if USE_PG:
            import psycopg2
            conn = psycopg2.connect(DATABASE_URL, sslmode='require')
            c = conn.cursor()
            c.execute(f"SELECT * FROM {table}")
            cols = [desc[0] for desc in c.description]
            rows = [dict(zip(cols, row)) for row in c.fetchall()]
            conn.close()
            return rows
        else:
            conn = sqlite3.connect(DB)
            conn.row_factory = sqlite3.Row
            c = conn.cursor()
            c.execute(f"SELECT * FROM {table}")
            rows = [dict(row) for row in c.fetchall()]
            conn.close()
            return rows
    except:
        return []

@app.route('/admin/backup')
@admin_required
def admin_backup():
    data = {}
    for table in ['users','attendance','leaves','leave_requests','notifications','shifts','urgent_reports','circulars','history_records','security_alerts','shifts_archive']:
        data[table] = _backup_rows(table)
    return send_file(
        io.BytesIO(json.dumps(data, ensure_ascii=False, default=str).encode('utf-8')),
        mimetype='application/json',
        as_attachment=True,
        download_name=f'backup_{ksa().strftime("%Y%m%d_%H%M")}.json'
    )

@app.route('/admin/restore', methods=['GET','POST'])
@admin_required
def admin_restore():
    if request.method == 'POST':
        file = request.files.get('file')
        if not file: flash('الرجاء اختيار ملف', 'danger'); return redirect(url_for('admin_restore'))
        try:
            data = json.load(file.stream)
        except: flash('الملف غير صالح', 'danger'); return redirect(url_for('admin_restore'))
        tables_map = {
            'users': ('id','username','password','full_name','role','chat_id','device_uid','is_blocked','phone_number','rank_title'),
            'attendance': ('id','user_id','action','latitude','longitude','timestamp','note'),
            'leaves': ('id','user_id','start_time','end_time','duration_label'),
            'leave_requests': ('id','user_id','duration_label','hours_duration','request_date','status'),
            'notifications': ('id','user_id','content','is_read','created_at'),
            'shifts': ('id','shift_date','current_duty'),
            'urgent_reports': ('id','sender_id','report_text','reply_text','created_at','status'),
            'circulars': ('id','content','created_at'),
            'history_records': ('id','user_id','note','created_at'),
            'security_alerts': ('id','user_id','alert_type','detail','created_at'),
            'shifts_archive': ('id','shift_date','current_duty','username','created_at'),
        }
        counts = {}
        for table, cols in tables_map.items():
            rows = data.get(table, [])
            if rows:
                placeholders = ','.join(['?' for _ in cols])
                cols_list = ','.join(cols)
                for row in rows:
                    try:
                        vals = [row.get(c) for c in cols]
                        if USE_PG:
                            db_run(f"INSERT INTO {table} ({cols_list}) VALUES ({placeholders}) ON CONFLICT DO NOTHING", vals)
                        else:
                            db_run(f"INSERT OR IGNORE INTO {table} ({cols_list}) VALUES ({placeholders})", vals)
                    except: pass
            counts[table] = len(rows)
        threading.Thread(target=backup.do_backup, daemon=True).start()
        flash(f'تم الاستيراد: {json.dumps(counts, ensure_ascii=False)}', 'success')
        return redirect(url_for('admin_dashboard'))
    return render_template('admin/restore.html')

init_db()
init_settings()
try: db_run("ALTER TABLE users ADD COLUMN avatar TEXT DEFAULT ''")
except: pass
try: db_run("ALTER TABLE users ADD COLUMN is_deleted INTEGER DEFAULT 0")
except: pass
try:
    if USE_PG:
        db_run("CREATE TABLE IF NOT EXISTS audit_log (id SERIAL PRIMARY KEY, admin_id INTEGER NOT NULL, action TEXT NOT NULL, target_id INTEGER, details TEXT, created_at TEXT NOT NULL)")
    else:
        db_run("CREATE TABLE IF NOT EXISTS audit_log (id INTEGER PRIMARY KEY AUTOINCREMENT, admin_id INTEGER NOT NULL, action TEXT NOT NULL, target_id INTEGER, details TEXT, created_at TEXT NOT NULL)")
except: pass
try: db_run("ALTER TABLE leaves ADD COLUMN is_active INTEGER DEFAULT 0")
except: pass
try: db_run("UPDATE leaves SET is_active=1")
except: pass
ADMIN_PASS = os.getenv('ADMIN_PASSWORD', '1000')
if not get_user('admn'):
    add_user('admn', ADMIN_PASS, 'القائد العام', 'admin')
    set_admin_phone('admn', os.getenv('ADMIN_PHONE', '0503077519'))

backup.do_restore()
backup.start_auto_backup()
clear_old_audit()

def _auto_save():
    threading.Thread(target=backup.do_backup, daemon=True).start()

@app.route('/admin/audit-log')
@admin_required
def admin_audit_log():
    logs = get_audit_log()
    return render_template('admin/audit_log.html', logs=logs)

@app.route('/admin/force-backup')
@admin_required
def admin_force_backup():
    _auto_save()
    flash('تم حفظ نسخة احتياطية', 'success')
    return redirect(url_for('admin_dashboard'))

if __name__ == '__main__':
    port = int(os.getenv('PORT', 5000))
    debug = os.getenv('FLASK_DEBUG', '0') == '1'
    print(f"المنظومة تعمل على المنفذ {port}")
    app.run(host='0.0.0.0', port=port, debug=debug)
